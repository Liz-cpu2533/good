import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# ------------------- 请修改以下参数 -------------------
CSV_FILE = "C:/Users/admin/Desktop/需求一选择题output.csv"  # 替换为你的 CSV 文件路径
ANSWER_COLUMN = "答案"  # 替换为答案列的列名
MODEL_ANSWER_COLUMN = "模型答案"  # 替换为模型答案列的列名
OUTPUT_FILE = None  # 输出文件名（可选，默认为原文件名加"_对比结果.xlsx"）
# ---------------------------------------------------

# 读取 CSV 文件
try:
    df = pd.read_csv(CSV_FILE)
except Exception as e:
    print(f"读取 CSV 文件时出错: {e}")
    exit()

# 检查列名是否存在
if ANSWER_COLUMN not in df.columns or MODEL_ANSWER_COLUMN not in df.columns:
    print(f"错误: 指定的列名不存在。CSV 文件中的列名有: {', '.join(df.columns)}")
    exit()

# 如果没有指定输出文件名，生成默认文件名
if OUTPUT_FILE is None:
    OUTPUT_FILE = CSV_FILE.rsplit('.', 1)[0] + '_对比结果.xlsx'

# 将 DataFrame 保存为 Excel 文件
df.to_excel(OUTPUT_FILE, index=False)

# 加载工作簿进行格式设置
wb = load_workbook(OUTPUT_FILE)
ws = wb.active

# 获取两列的列索引（转换为 Excel 列字母）
answer_col_letter = get_column_letter(df.columns.get_loc(ANSWER_COLUMN) + 1)
model_answer_col_letter = get_column_letter(df.columns.get_loc(MODEL_ANSWER_COLUMN) + 1)

# 创建黄色填充样式
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# 从第2行开始（第1行是表头）
for row_num in range(2, len(df) + 2):
    # 获取两列的值
    answer_value = ws[f'{answer_col_letter}{row_num}'].value
    model_answer_value = ws[f'{model_answer_col_letter}{row_num}'].value

    # 处理 NaN 值
    if pd.isna(answer_value) and pd.isna(model_answer_value):
        is_different = False
    elif pd.isna(answer_value) or pd.isna(model_answer_value):
        is_different = True
    else:
        # 直接比较值（忽略首尾空格）
        is_different = str(answer_value).strip() != str(model_answer_value).strip()

    # 如果不同，整行标黄
    if is_different:
        for cell in ws[row_num]:
            cell.fill = yellow_fill

# 保存修改后的工作簿
wb.save(OUTPUT_FILE)
print(f"对比完成，结果已保存到: {OUTPUT_FILE}")